from flask import Flask, render_template,request,redirect
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///employer.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config['SECRET_KEY'] = 'login_system'
db = SQLAlchemy(app)

class Employee(db.Model):
    Sno = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Emp_id = db.Column(db.String(500))
    Name = db.Column(db.String(500))
    Designation = db.Column(db.String(500))
    Department = db.Column(db.String(500))
    Project = db.Column(db.String(500))
    Job_role = db.Column(db.String(500))
    Employment_status = db.Column(db.String(500))
    Joining_date = db.Column(db.String(500))
    Experience = db.Column(db.String(500))
    Location = db.Column(db.String(500))
    Last_promoted = db.Column(db.String(500))
    Comments = db.Column(db.String(500))

def extract_data_from_excel():
    wb = load_workbook("employee_data 1.xlsx")
    ws = wb.active
    column_mappings = {
        'Sno': 0,
        'Emp_id': 1,
        'Name': 2,
        'Designation': 3,
        'Department': 4,
        'Project': 5,
        'Job_role': 6,
        'Employment_status': 7,
        'Joining_date': 8,
        'Experience': 9,
        'Location': 10,
        'Last_promoted': 11,
        'Comments': 12
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        
        if not all(cell is None for cell in row):
            Sno = row[column_mappings['Sno']]
            Emp_id = row[column_mappings['Emp_id']]
            Name = row[column_mappings['Name']]
            Designation = row[column_mappings['Designation']]
            Department = row[column_mappings['Department']]
            Project = row[column_mappings['Project']]
            Job_role = row[column_mappings['Job_role']]
            Employment_status = row[column_mappings['Employment_status']]
            Joining_date = row[column_mappings['Joining_date']]
            Experience = row[column_mappings['Experience']]
            Location = row[column_mappings['Location']]
            Last_promoted = row[column_mappings['Last_promoted']]
            Comments = row[column_mappings['Comments']]

            existing_data = Employee.query.filter_by(Name=Name).first()
            if not existing_data:
                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                    Department=Department, Project=Project, Job_role=Job_role,
                                    Employment_status=Employment_status, Joining_date=Joining_date,
                                    Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                    Comments=Comments)
                db.session.add(employee)
    db.session.commit()

@app.route("/")
def Home():
    data=Employee.query.all()
    return render_template("index.html",data=data)

@app.route("/add", methods=["GET", "POST"])
def Add():
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        experience = request.form.get("experience")
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        existing_data=Employee.query.filter_by(Name=name).first()
        if not existing_data:
            employee = Employee(
                Emp_id=emp_id,
                Name=name,
                Designation=designation,
                Department=department,
                Project=project,
                Job_role=job_role,
                Employment_status=employment_status,
                Joining_date=joining_date,
                Experience=experience,
                Location=location,
                Last_promoted=last_promoted,
                Comments=comments
            )
            db.session.add(employee)
            db.session.commit()
        return redirect("/")
    return render_template("add.html")

@app.route("/update/<int:sno>",methods=["GET","POST"])
def Update(sno):
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        experience = request.form.get("experience")
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee=Employee.query.filter_by(Sno=sno).first()
        employee.Emp_id=emp_id
        employee.Name=name
        employee.Designation=designation
        employee.Department=department
        employee.Project=project
        employee.Job_role=job_role
        employee.Employment_status=employment_status
        employee.Joining_date=joining_date
        employee.Experience=experience
        employee.Location=location
        employee.Last_promoted=last_promoted
        employee.Comments=comments
        db.session.add(employee)
        db.session.commit()
        return redirect("/")
    employee=Employee.query.filter_by(Sno=sno).first()
    return render_template("update.html",employee=employee)

@app.route("/delete/<int:sno>")
def Delete(sno):
    employee=Employee.query.filter_by(Sno=sno).first()
    db.session.delete(employee)
    db.session.commit()
    return redirect("/")
with app.app_context():
        db.create_all()
        data=extract_data_from_excel()
         
if "__name__" == "__main__":
    
    app.run(debug=True)   



